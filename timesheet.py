import os
import sys
import json
import locale
from datetime import datetime
from datetime import timedelta
from openpyxl import load_workbook

class Timesheet(object):
    """Object for managing work hours timesheet.
    """
    def __init__(self):
        """Constructor,  initializes 'config' instance attribute.

        :param str config_file: Name of configuration file
        """
        self.config_file = os.path.basename(__file__).split(".")[0] + ".json"
        program_path = os.path.dirname(os.path.realpath(__file__))
        config = self.load_json_file(self.config_file)
        if config == None:
            sys.exit("Exiting. Configuration file '{}' not found.".format(self.config_file))
        self.config = {}
        if config["name"]:
            self.config["name"] = config["name"]
        else:
            self.config["name"] = ""
        # set records configuration
        self.config["records"] = {}
        self.config["records"]["records_name"] = "timesheet_{}_{}.json"
        if config["records"] and config["records"]["records_dir"]:
            self.config["records"]["records_dir"] = config["records"]["records_dir"]
        else:
            self.config["records"]["records_dir"] = os.path.join(program_path, "data/records")
        # set exports configuration
        self.config["exports"] = {}
        self.config["exports"]["exports_name"] = "timesheet_{}_{}.xlsx"
        if config["exports"] and config["exports"]["exports_dir"]:
            self.config["exports"]["exports_dir"] = config["exports"]["exports_dir"]
        else:
            self.config["exports"]["exports_dir"] = os.path.join(program_path, "data/exports")
        # set templates configuration
        self.config["templates"] = {}
        self.config["templates"]["templates_name"] = "template_timesheet_{}_days.xlsx"
        self.config["templates"]["templates_dir"] = os.path.join(program_path, "data/templates")
        if not os.path.isdir(self.config["records"]["records_dir"]):
            os.makedirs(self.config["records"]["records_dir"])
        if not os.path.isdir(self.config["exports"]["exports_dir"]):
            os.makedirs(self.config["exports"]["exports_dir"])

    def list_config(self):
        return self.config

    def set_config(self, key, value):
        self.config[key] = value
        self.write_json_file(self.config_file, self.config)

    def load_records(self, date):
        """Initializes 'records_file' and 'records' instance attributes.

        :param datetime.date date: Date of the records that are to be loaded,
            only year and month are relevant
        """
        self.records_file = os.path.join(
            self.config["records"]["records_dir"],
            self.config["records"]["records_name"].format(date.year, date.strftime("%m"))
        )
        self.records = self.load_json_file(self.records_file, [])

    def netto_workdays(self, start_date, end_date, holidays=[], weekend_days=[5,6]):
        """Calculates number of workdays between two given dates, subtracting weekends.

        :param date start_date: Date from where to start counting
        :param date end_date: Date where to stop counting
        :param list holidays: List of holidays, date objects
        :param list weekend_days: List of days included in weekend; 5=sat, 6=sun
        :return: Integer number of workdays
        """
        delta_days = (end_date - start_date).days + 1
        full_weeks, extra_days = divmod(delta_days, 7)
        # num_workdays = how many days/week you work * total number of weeks
        num_workdays = (full_weeks + 1) * (7 - len(weekend_days))
        # subtract out any working days that fall in the 'shortened week'
        for d in range(1, 8 - extra_days):
            if (end_date + timedelta(d)).weekday() not in weekend_days:
                 num_workdays -= 1
        # skip holidays that fall on weekend_days
        holidays =  [x for x in holidays if x.weekday() not in weekend_days]
        # subtract out any holidays
        for d in holidays:
            if start_date <= d <= end_date:
                num_workdays -= 1
        return num_workdays

    def add_record(self, date, work_hours, break_time, comment, special):
        """Add a new record in timesheet.

        :param datetime.date date: Date of record
        :param tuple work_hours: Two datetime.time objects representing start and end of workday
        :param tuple break_time: Two datetime.time objects representing start and end of break time
        :param str comment: Comment for the record
        :param bool special: Whether the record is special or not
        :rtype: bool
        """
        self.load_records(date)
        if not self.record_exists(date):
            record = self.create_record(date, work_hours, break_time, comment, special)
            self.records.append(record)
            self.write_json_file(self.records_file, self.records)
            return True
        return False

    def delete_record(self, date):
        """Delete a record from timesheet.

        :param datetime.date date: Date of record
        :rtype: bool
        """
        self.load_records(date)
        for record in self.records:
            if date.strftime("%d.%m.%Y") == record["date"]:
                self.records.remove(record)
                if len(self.records) > 0:
                    self.write_json_file(self.records_file, self.records)
                else:
                    os.remove(self.records_file)
                return True
        return False

    def update_record(self, date, work_hours, break_time, comment, special):
        """Replace a record in timesheet.

        :param datetime.date date: Date of record
        :param tuple work_hours: Two datetime.time objects representing start and end of workday
        :param tuple break_time: Two datetime.time objects representing start and end of break time
        :param str comment: Comment for the record
        :param bool special: Whether the record is special or not
        :rtype: bool
        """
        self.load_records(date)
        new_record = self.create_record(date, work_hours, break_time, comment, special)
        for record in self.records:
            if date.strftime("%d.%m.%Y") == record["date"] and not record == new_record:
                record.update(new_record)
                self.write_json_file(self.records_file, self.records)
                return True
        return False

    def create_record(self, date, work_hours, break_time, comment, special):
        """Create a record as dictionary.

        :param datetime.date date: Date of record
        :param tuple work_hours: Two datetime.time objects representing start and end of workday
        :param tuple break_time: Two datetime.time objects representing start and end of break time
        :param str comment: Comment for the record
        :param bool special: Whether the record is special or not
        :return: Dictionary with record data
        """
        return {
            "date": date.strftime("%d.%m.%Y"),
            "start_day": work_hours[0].strftime("%H:%M"),
            "end_day": work_hours[1].strftime("%H:%M"),
            "start_break": break_time[0].strftime("%H:%M"),
            "end_break": break_time[1].strftime("%H:%M"),
            "comment": comment,
            "special": str(special)
        }

    def write_json_file(self, file, content):
        """Write list of records to JSON file.

        :param str file: Name of file to write
        :param content: Content to write in file
        """
        with open(file, "w", encoding="utf-8") as f:
            json.dump(content, f, indent=2)

    def load_json_file(self, file, default_content=None):
        """Load JSON file, return content.

        :param str file: Name of file to load
        :param default_content: Default content to return as loaded content
        :return: Content from file
        """
        if os.path.isfile(file) and os.path.getsize(file):
            with open(file, "r", encoding="utf-8") as f:
                return json.load(f)
        return default_content

    def record_exists(self, date):
        """Check if record exists already.

        :param datetime.date: Date to look for
        return: Bool with the result
        """
        for record in self.records:
            if date.strftime("%d.%m.%Y") == record["date"]:
                return True
        return False

    def extract_carryover_hours(self, file):
        """Reads excel file, returns carryover hours.

        :param str file: name of file to read
        return: number of carryover hours from file
        """
        try:
            wb = load_workbook(file, data_only=True)
        except FileNotFoundError:
            return 0
        row = wb["Logging"].cell(row=2, column=2).value
        column = wb["Logging"].cell(row=3, column=2).value
        return wb["Timesheet"].cell(row=row, column=column).value

    def export(self, date):
        """Export timesheet as .xlsx file

        :param datetime.date date:  Date of the timesheet to be exported
        """
        self.load_records(date)
        if len(self.records) == 0:
            exit_message = "Exiting. There are no records for {} {} to export.".format(date.strftime("%B"), date.year)
            sys.exit(exit_message)
        if not self.config["name"]:
            print("Warning. Your name is missing from the configuration.")

        total_days = (date.replace(month = date.month % 12 +1, day = 1)-timedelta(days=1)).day
        start_month = date.replace(day = 1)
        end_month = date.replace(day = total_days)
        workdays = self.netto_workdays(start_month, end_month, weekend_days=(5,6))
        template_file = os.path.join(self.config["templates"]["templates_dir"],
            self.config["templates"]["templates_name"].format(workdays))

        export_file = os.path.join(
            self.config["exports"]["exports_dir"],
            self.config["exports"]["exports_name"].format(date.year, date.strftime("%m"))
        )
        prev_month = (start_month - timedelta(days=1)).strftime("%m")
        prev_export_file = os.path.join(
            self.config["exports"]["exports_dir"],
            self.config["exports"]["exports_name"].format(date.year, prev_month)
        )
        carryover_hours = self.extract_carryover_hours(prev_export_file)

        # set locale to use weekdays, months full name in german
        locale.setlocale(locale.LC_TIME, 'de_DE.UTF-8')
        wb = load_workbook(template_file)
        ws = wb["Timesheet"]
        ws.cell(row=7, column=4).value = self.config["name"]
        month_year_str = "{} {}".format(date.strftime("%B"), date.year)
        ws.cell(row=8, column=4).value = month_year_str
        ws.cell(row=8, column=10).value = carryover_hours
        row = 12
        for record in self.records:
            col = 2
            record_date =  datetime.strptime(record["date"], "%d.%m.%Y")
            ws.cell(row=row, column=col).value = record_date.strftime("%A")
            col += 1
            ws.cell(row=row, column=col).value = record_date
            col += 1
            if "special" in record.keys() and record["special"] == "True":
                ws.cell(row=row, column=9).value = 8.00
                col += 4
            else:
                ws.cell(row=row, column=col).value = datetime.strptime(record["start_day"], "%H:%M").time()
                col += 1
                ws.cell(row=row, column=col).value = datetime.strptime(record["end_day"], "%H:%M").time()
                col += 1
                ws.cell(row=row, column=col).value = datetime.strptime(record["start_break"], "%H:%M").time()
                col += 1
                ws.cell(row=row, column=col).value = datetime.strptime(record["end_break"], "%H:%M").time()
            col += 4
            ws.cell(row=row, column=col).value = record["comment"]
            row += 1
        wb.save(export_file)
        return True
